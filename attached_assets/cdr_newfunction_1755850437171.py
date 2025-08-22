#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cdr_gui_app.py — Final

What’s in here:
• Multi-file CDR ingest (CSV), robust normalization
• 16 Excel sheets (Mapping, Summary, MaxCalls, MaxDuration, MaxStay, OtherStateContactSummary,
  RoamingPeriod, IMEIPeriod, IMSIPeriod, Night_Mapping, Night_MaxStay, Day_Mapping, Day_MaxStay,
  WorkHomeLocation, HomeLocationBasedonDayFirstand, ISDCalls)
• Fix: Party B for SMS preserves sender code (e.g., "VK-NSESMS") instead of replacing with a number
• Night window is 18:00–06:00 for Night_* sheets
• IMEIPeriod guarantees unique IMEI rows per CdrNo (keeps highest activity)
• Excel output: AutoFilter + Freeze top row, auto-fit column widths, and trims spaces in text
"""

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import pandas as pd
import numpy as np
import re, os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
# Night hours inclusive: hour >= 18  OR hour < 6
NIGHT_START = 18
NIGHT_END = 6

# alias map for common column name variants
ALIASES = {
    "target_number": ["target /a party number","target no","cdr party no","target"],
    "a_party": ["calling party telephone number","a party number","msisdn a","a_number","calling party"],
    "b_party": ["called party telephone number","b party number","b party no","called party"],
    "call_forwarding": ["call forwarding","call fow no","call forwarding number","callforward"],
    "lrn_called": ["lrn called no","lrn no","translation of lrn","lrn"],
    "call_date": ["call date","date"],
    "call_time": ["call time","time","call initiation time","start time"],
    "call_term_time": ["call termination time","end time","release time"],
    "duration": ["call duration","dur(s)","duration","hold time (sec)","durations"],
    "call_type": ["call type","service type","event type","dir","toc","type of connection"],
    "toc": ["toc","type of connection"],
    "first_cell_id": ["first cell id","first cgi","first cell global id","firstcellid"],
    "last_cell_id": ["last cell id","last cgi","last cell global id","lastcellid"],
    "first_cell_addr": ["first bts location","first cell site address","first cell address","first cell site location"],
    "last_cell_addr": ["last bts location","last cell site address","last cell address","last cell site location"],
    "first_cell_city": ["first cell site name-city","first cell site name","first cell city","first site name"],
    "last_cell_city": ["last cell site name-city","last cell site name","last cell city","last site name"],
    "first_lat_long": ["first lat/long","first cgi lat/long","first latitude/longitude","first lat long"],
    "last_lat_long": ["last lat/long","last cgi lat/long","last latitude/longitude","last lat long"],
    "smsc": ["sms center number","smsc no","smsc","sms center"],
    "imei": ["imei","esn_imei_a","device imei"],
    "imsi": ["imsi","imsi_a","subscriber imsi"],
    "circle": ["roaming circle name","roam circle","circle","roaming"],
    "home_circle": ["home circle","home region","home state"],
    "operator": ["operator","service provider","sp","provider"]
}

# ---------------- Helpers ----------------
def _lower_map(cols):
    seen = {}
    for c in cols:
        cl = str(c).lower().strip()
        if cl not in seen:
            seen[cl] = c
    return seen

def _pick(cols_map, df, candidates):
    for c in candidates:
        if c in cols_map:
            return df[cols_map[c]]
    return pd.Series([np.nan]*len(df))

def detect_header_start(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        for i, line in enumerate(lines):
            l = line.lower()
            if ("calling party telephone number" in l) or ("target /a party number" in l) or ("target no" in l):
                return i
    except Exception:
        pass
    return 0

def load_csv_like_raw(path):
    start = detect_header_start(path)
    df = pd.read_csv(
        path, engine="python", sep=",", header=0, skiprows=start,
        on_bad_lines="skip", dtype=str
    )
    df = df.loc[:, ~pd.Index(df.columns).duplicated()]
    return df

def to_seconds(x):
    if pd.isna(x): return 0
    s = str(x).strip().strip("'")
    if s == "": return 0
    if s.isdigit():
        return int(s)
    try:
        parts = s.split(":")
        if len(parts) == 3:
            h,m,ss = map(int, parts); return h*3600 + m*60 + ss
        if len(parts) == 2:
            m,ss = map(int, parts); return m*60 + ss
    except Exception:
        pass
    try:
        return int(float(s))
    except Exception:
        return 0

def parse_time_field(x):
    s = str(x).strip().strip("'")
    for fmt in ("%H:%M:%S","%H:%M","%I:%M:%S %p","%I:%M %p"):
        try:
            return pd.to_datetime(s, format=fmt).time()
        except Exception:
            continue
    if re.fullmatch(r"\d{6}", s):
        try:
            return pd.to_datetime(s, format="%H%M%S").time()
        except:
            return None
    if re.fullmatch(r"\d{4}", s):
        try:
            return pd.to_datetime(s, format="%H%M").time()
        except:
            return None
    return None

def parse_date_field(x):
    s = str(x).strip().strip("'").replace(".", "/")
    for dayfirst in (True, False):
        try:
            return pd.to_datetime(s, dayfirst=dayfirst, errors="raise").date()
        except Exception:
            continue
    return None

def fmt_date_dMonY(d):
    if pd.isna(d) or d is None:
        return ""
    try:
        return pd.to_datetime(d).strftime("%d/%b/%Y")
    except Exception:
        return ""

def fmt_time_HMS(t):
    if pd.isna(t) or t is None:
        return ""
    try:
        import datetime as _dt
        if isinstance(t, _dt.time):
            return t.strftime("%H:%M:%S")
        if isinstance(t, _dt.date) and not isinstance(t, _dt.datetime):
            return ""
        return pd.to_datetime(t).strftime("%H:%M:%S")
    except Exception:
        return ""

def normalize_msisdn(num):
    s = re.sub(r"\D", "", str(num))
    if s.startswith("0"):
        s = s[1:]
    if len(s) > 10 and s.startswith("91"):
        s = s[2:]
    return s

def contains_sender_code(s):
    """True if looks like an alphanumeric sender-id (e.g., VK-NSESMS, AD-ICICIB)."""
    if s is None: return False
    st = str(s).strip()
    if st == "": return False
    # sender-id typically contains letters, optionally hyphen/spaces
    return bool(re.search(r"[A-Za-z]", st))

def clean_text(s):
    if s is None or (isinstance(s,float) and np.isnan(s)): return ""
    return re.sub(r"\s+", " ", str(s)).strip()

def is_night_hour(hour):
    if pd.isna(hour): return False
    # 18..23 or 0..5
    return (hour >= NIGHT_START) or (hour < NIGHT_END)

def safe_reindex_columns(df, columns):
    df = df.loc[:, ~df.columns.duplicated()]
    cols = []
    seen = set()
    for c in columns:
        if c not in seen:
            cols.append(c); seen.add(c)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df.reindex(columns=cols)

# ---------------- STANDARDIZE ----------------
def standardize_rows(df):
    cols = _lower_map(df.columns)
    pick = lambda k: _pick(cols, df, ALIASES[k])

    std = pd.DataFrame({
        'TargetRaw': pick('target_number').astype(str),
        'Araw': pick('a_party').astype(str),
        'Braw': pick('b_party').astype(str),
        'CallForward': pick('call_forwarding').astype(str),
        'LRN': pick('lrn_called').astype(str),
        'CallDateRaw': pick('call_date').astype(str),
        'CallTimeRaw': pick('call_time').astype(str),
        'CallTermTimeRaw': pick('call_term_time').astype(str),
        'DurationRaw': pick('duration').astype(str),
        'CallTypeRaw': pick('call_type').astype(str),
        'TOC': pick('toc').astype(str),
        'FirstCellID': pick('first_cell_id').astype(str),
        'LastCellID': pick('last_cell_id').astype(str),
        'FirstCellAddr': pick('first_cell_addr').astype(str),
        'LastCellAddr': pick('last_cell_addr').astype(str),
        'FirstCellCity': pick('first_cell_city').astype(str),
        'LastCellCity': pick('last_cell_city').astype(str),
        'IMEI': pick('imei').astype(str),
        'IMSI': pick('imsi').astype(str),
        'Circle': pick('circle').astype(str),
        'HomeCircle': pick('home_circle').astype(str),
        'operator': pick('operator').astype(str),
        'SMSC': _pick(cols, df, ALIASES.get('smsc', []))
    })

    std['DateObj'] = std['CallDateRaw'].apply(parse_date_field)
    std['TimeObj'] = std['CallTimeRaw'].apply(parse_time_field)
    std['start_dt'] = pd.to_datetime(
        [pd.NaT if (d is None or t is None) else f"{d} {t}" for d,t in zip(std['DateObj'], std['TimeObj'])],
        errors="coerce"
    )
    std['Duration'] = std['DurationRaw'].apply(to_seconds).astype('Int64')

    def derive_call_type(ct, toc):
        s = f"{ct} {toc}".lower()
        is_sms = 'sms' in s
        # Detect inbound/outbound more robustly, including patterns like 'smsin', 'sms_in', 'mo', 'mt'
        if is_sms:
            if 'smsin' in s or 'sms_in' in s or 'inbound' in s or 'terminat' in s or 'mt' in s:
                d = 'IN'
            elif 'smsout' in s or 'sms_out' in s or 'mo' in s or 'orig' in s:
                d = 'OUT'
            else:
                # fallback: if string contains ' in' or endswith 'in'
                d = 'IN' if (' in' in s or s.strip().endswith('in')) else 'OUT'
        else:
            if any(k in s for k in ['incoming','mt','terminating','a_in','term',' in','-in']):
                d = 'IN'
            elif any(k in s for k in ['outgoing','mo','originating','a_out','orig',' out','-out']):
                d = 'OUT'
            else:
                d = 'OUT'
        return ('SMS' if is_sms else 'CALL') + '_' + d
    std['CallTypeStd'] = [derive_call_type(ct,toc) for ct,toc in zip(std['CallTypeRaw'], std['TOC'])]

    # Normalized numbers (for calls)
    std['A_norm'] = std['Araw'].apply(normalize_msisdn).astype(str)
    std['B_norm'] = std['Braw'].apply(normalize_msisdn).astype(str)
    std['Target_norm'] = std['TargetRaw'].apply(lambda x: normalize_msisdn(x) if pd.notna(x) else "")

    # Determine monitored number (CdrNo)
    if std['Target_norm'].str.strip().replace('', np.nan).dropna().shape[0] > 0:
        try:
            top_target = std['Target_norm'].replace('', np.nan).mode().iat[0]
        except Exception:
            top_target = ""
    else:
        combined = pd.concat([std['A_norm'], std['B_norm']], ignore_index=True)
        combined = combined[combined.astype(str) != ""]
        try:
            top_target = combined.mode().iat[0]
        except Exception:
            top_target = ""
    top_target = str(top_target) if top_target is not None else ""
    std['CdrNo'] = top_target

    # Counterparty selection with SMS sender-code logic
    def pick_counterparty(row):
        a_raw = clean_text(row['Araw'])
        b_raw = clean_text(row['Braw'])
        a_num = str(row['A_norm'])
        b_num = str(row['B_norm'])
        t = str(row['CdrNo'])
        ctype = str(row['CallTypeStd'])

        # If SMS and any side has an alphanumeric sender-id, use that as "B Party"
        if ctype.startswith("SMS"):
            if contains_sender_code(b_raw):
                return b_raw
            if contains_sender_code(a_raw):
                return a_raw

        # Otherwise choose the opposite number of target (call cases)
        if a_num == t and b_num:
            return b_num
        if b_num == t and a_num:
            return a_num
        if b_num:
            return b_num
        if a_num:
            return a_num
        # Fall back to raw if everything empty
        return b_raw or a_raw or ""

    std['Counterparty'] = std.apply(pick_counterparty, axis=1).astype(str)

    # Helpful derived fields
    std['Hour'] = std['start_dt'].dt.hour
    std['IsNight'] = std['Hour'].apply(is_night_hour)
    std['DateStr'] = std['start_dt'].dt.date.apply(fmt_date_dMonY)
    std['TimeStr'] = std['start_dt'].dt.time.apply(fmt_time_HMS)

    # Clean IMEI & IMSI strings consistently
    std['IMEI'] = std['IMEI'].apply(clean_text)
    std['IMSI'] = std['IMSI'].apply(clean_text)
    std['operator'] = std['operator'].astype(str).fillna("")

    # Drop rows where Counterparty is truly blank
    std = std[std['Counterparty'].astype(str).str.strip() != ""].reset_index(drop=True)
    return std

# ---------------- Sheet builders with exact headers ----------------

def empty_series(n, val=""):
    return pd.Series([val]*n)

# 1) Mapping
MAP_COLUMNS = [
    'CdrNo','B Party','Date','Time','Duration','Call Type','First Cell ID',
    'First Cell ID Address','Last Cell ID','Last Cell ID Address','IMEI',
    'IMEI Manufacturer','Device Type','IMSI','Roaming','B Party Provider',
    'Main City(First CellID)','Sub City(First CellID)','Lat-Long-Azimuth (First CellID)',
    'Crime','Circle','Operator','CallForward','LRN','Location'
]

def build_mapping(std, include_location=False):
    std = std.reset_index(drop=True)
    n = len(std)
    df = pd.DataFrame({
        'CdrNo': std['CdrNo'],
        'B Party': std['Counterparty'].apply(clean_text),
        'Date': std['DateStr'],
        'Time': std['TimeStr'],
        'Duration': std['Duration'].fillna(0).astype('Int64'),
        'Call Type': std['CallTypeStd'],
        'First Cell ID': std['FirstCellID'].apply(clean_text),
        'First Cell ID Address': std['FirstCellAddr'].apply(clean_text),
        'Last Cell ID': std['LastCellID'].apply(clean_text),
        'Last Cell ID Address': std['LastCellAddr'].apply(clean_text),
        'IMEI': std['IMEI'],
        'IMEI Manufacturer': empty_series(n, ""),
        'Device Type': empty_series(n, ""),
        'IMSI': std['IMSI'],
        'Roaming': std['Circle'].fillna("").astype(str),
        'B Party Provider': empty_series(n, ""),
        'Main City(First CellID)': std['FirstCellCity'].apply(clean_text),
        'Sub City(First CellID)': empty_series(n, ""),
        'Lat-Long-Azimuth (First CellID)': empty_series(n, ""),
        'Crime': empty_series(n, ""),
        'Circle': std['Circle'].fillna("").astype(str),
        'Operator': std['operator'].fillna("").astype(str),
        'CallForward': std['CallForward'].apply(clean_text),
        'LRN': std['LRN'].apply(clean_text),
        'Location': empty_series(n, "") if include_location else empty_series(n, "")
    })
    return safe_reindex_columns(df, MAP_COLUMNS)

# 2) Summary
SUM_COLUMNS = [
 'CdrNo','B Party','Provider','Type','Total Calls','Out Calls','In Calls',
 'Out Sms','In Sms','Other Calls','Roam Calls','Roam Sms','Total Duration',
 'Total Days','Total CellIds','Total Imei','Total Imsi',
 'First Call Date','First Call Time','Last Call Date','Last Call Time'
]

def build_summary(std):
    std = std.reset_index(drop=True)
    std2 = std[std['Counterparty'].astype(str).str.strip() != ""].copy()
    grp = std2.groupby(['CdrNo','Counterparty'], dropna=False)
    agg = grp.agg(
        Total_Calls=('Counterparty','count'),
        Total_Duration=('Duration','sum'),
        First_dt=('start_dt','min'),
        Last_dt=('start_dt','max'),
        Total_Days=('DateStr', lambda s: s.nunique()),
        Total_CellIds=('FirstCellID', lambda s: s.nunique()),
        Total_IMEI=('IMEI', lambda s: s.nunique()),
        Total_IMSI=('IMSI', lambda s: s.nunique()),
        OutCalls=('CallTypeStd', lambda s: (s=='CALL_OUT').sum()),
        InCalls=('CallTypeStd', lambda s: (s=='CALL_IN').sum()),
        OutSms=('CallTypeStd', lambda s: (s=='SMS_OUT').sum()),
        InSms=('CallTypeStd', lambda s: (s=='SMS_IN').sum()),
        RoamCalls=('Circle', lambda s: s.replace("", np.nan).notna().sum()),
        RoamSms=('CallTypeStd', lambda s: 0)
    ).reset_index()

    # provider: most frequent operator per counterparty
    operator_map = {}
    for cp, sub in std2.groupby('Counterparty'):
        try:
            operator_map[cp] = sub['operator'].astype(str).mode().iat[0]
        except Exception:
            operator_map[cp] = ""

    df = pd.DataFrame({
        'CdrNo': agg['CdrNo'],
        'B Party': agg['Counterparty'].apply(clean_text),
        'Provider': agg['Counterparty'].map(operator_map).fillna(""),
        'Type': "",
        'Total Calls': agg['Total_Calls'],
        'Out Calls': agg['OutCalls'],
        'In Calls': agg['InCalls'],
        'Out Sms': agg['OutSms'],
        'In Sms': agg['InSms'],
        'Other Calls': 0,
        'Roam Calls': agg['RoamCalls'],
        'Roam Sms': agg['RoamSms'],
        'Total Duration': agg['Total_Duration'],
        'Total Days': agg['Total_Days'],
        'Total CellIds': agg['Total_CellIds'],
        'Total Imei': agg['Total_IMEI'],
        'Total Imsi': agg['Total_IMSI'],
        'First Call Date': agg['First_dt'].dt.date.apply(fmt_date_dMonY),
        'First Call Time': agg['First_dt'].dt.time.apply(fmt_time_HMS),
        'Last Call Date': agg['Last_dt'].dt.date.apply(fmt_date_dMonY),
        'Last Call Time': agg['Last_dt'].dt.time.apply(fmt_time_HMS),
    })
    return safe_reindex_columns(df, SUM_COLUMNS)

# 3) MaxCalls
MAXC_COLUMNS = ['CdrNo','B Party','Total Calls','Provider']
def build_max_calls(std):
    std = std.reset_index(drop=True)
    gsrc = std[std['Counterparty'].astype(str).str.strip() != ""].copy()
    g = gsrc.groupby(['CdrNo','Counterparty'], dropna=False).size().reset_index(name='Total Calls')
    g = g[g['Counterparty'].astype(str).str.strip() != ""]
    # Provider from operator mode
    op_map = {}
    for cp, sub in gsrc.groupby('Counterparty'):
        try:
            op_map[cp] = sub['operator'].astype(str).mode().iat[0]
        except Exception:
            op_map[cp] = ""
    g['Provider'] = g['Counterparty'].map(op_map).fillna("")
    g = g.rename(columns={'Counterparty': 'B Party'})
    g = g[g['B Party'].astype(str).str.strip() != ""]
    return safe_reindex_columns(g, MAXC_COLUMNS)

# 4) MaxDuration
MAXD_COLUMNS = ['CdrNo','B Party','Total Duration','Provider']
def build_max_duration(std):
    std = std.reset_index(drop=True)
    gsrc = std[std['Counterparty'].astype(str).str.strip() != ""].copy()
    g = gsrc.groupby(['CdrNo','Counterparty'], dropna=False)['Duration'].sum().reset_index(name='Total Duration')
    g = g[g['Counterparty'].astype(str).str.strip() != ""]
    op_map = {}
    for cp, sub in gsrc.groupby('Counterparty'):
        try:
            op_map[cp] = sub['operator'].astype(str).mode().iat[0]
        except Exception:
            op_map[cp] = ""
    g['Provider'] = g['Counterparty'].map(op_map).fillna("")
    g = g.rename(columns={'Counterparty': 'B Party'})
    g = g[g['B Party'].astype(str).str.strip() != ""]
    return safe_reindex_columns(g, MAXD_COLUMNS)

# 5) MaxStay
STAY_COLUMNS = [
 'CdrNo','Cell ID','Total Calls','Days','Tower Address','Latitude','Longitude',
 'Azimuth','Roaming','First Call Date','First Call Time','Last Call Date','Last Call Time'
]
def build_max_stay(std):
    std = std.reset_index(drop=True)
    sub = std[std['FirstCellID'].astype(str).str.strip() != ""].copy()
    if sub.empty:
        return pd.DataFrame(columns=STAY_COLUMNS)
    g = sub.groupby(['CdrNo','FirstCellID'], dropna=False).agg(
        Total_Calls=('FirstCellID','count'),
        Days=('DateStr', lambda s: s.nunique()),
        TowerAddress=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else ""),
        First_dt=('start_dt','min'),
        Last_dt=('start_dt','max'),
        Roaming=('Circle', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else "")
    ).reset_index()
    df = pd.DataFrame({
        'CdrNo': g['CdrNo'],
        'Cell ID': g['FirstCellID'],
        'Total Calls': g['Total_Calls'],
        'Days': g['Days'],
        'Tower Address': g['TowerAddress'].apply(clean_text),
        'Latitude': "",
        'Longitude': "",
        'Azimuth': "",
        'Roaming': g['Roaming'],
        'First Call Date': g['First_dt'].dt.date.apply(fmt_date_dMonY),
        'First Call Time': g['First_dt'].dt.time.apply(fmt_time_HMS),
        'Last Call Date': g['Last_dt'].dt.date.apply(fmt_date_dMonY),
        'Last Call Time': g['Last_dt'].dt.time.apply(fmt_time_HMS),
    })
    return safe_reindex_columns(df, STAY_COLUMNS)

# 6) OtherStateContactSummary
OSCS_COLUMNS = ['CdrNo','Circle','Total Calls','Out Calls','In Calls','Out Sms','In Sms','Other Calls','Total Duration']
def build_other_state_contact_summary(std):
    std = std.reset_index(drop=True)
    sub = std.copy()
    sub['Circle'] = sub['Circle'].astype(str)
    sub = sub[sub['Circle'].str.strip() != ""]
    if sub.empty:
        return pd.DataFrame(columns=OSCS_COLUMNS)
    g = sub.groupby(['CdrNo','Circle'], dropna=False).agg(
        TotalCalls=('Circle','count'),
        OutCalls=('CallTypeStd', lambda s: (s=='CALL_OUT').sum()),
        InCalls=('CallTypeStd', lambda s: (s=='CALL_IN').sum()),
        OutSms=('CallTypeStd', lambda s: (s=='SMS_OUT').sum()),
        InSms=('CallTypeStd', lambda s: (s=='SMS_IN').sum()),
        TotalDuration=('Duration','sum')
    ).reset_index()
    df = pd.DataFrame({
        'CdrNo': g['CdrNo'],
        'Circle': g['Circle'],
        'Total Calls': g['TotalCalls'],
        'Out Calls': g['OutCalls'],
        'In Calls': g['InCalls'],
        'Out Sms': g['OutSms'],
        'In Sms': g['InSms'],
        'Other Calls': 0,
        'Total Duration': g['TotalDuration']
    })
    df = df[df['Total Calls'] > 0]
    return safe_reindex_columns(df, OSCS_COLUMNS)

# 7) RoamingPeriod
ROAM_COLUMNS = ['CdrNo','Roaming','Period','Total Calls','Days','First Location','Last Location',
 'Out Calls','In Calls','Out Sms','In Sms','Other Calls','Total Duration']
def build_roaming_period(std):
    std = std.reset_index(drop=True)
    sub = std.copy()
    sub['Circle'] = sub['Circle'].astype(str)
    sub = sub[sub['Circle'].str.strip() != ""]
    if sub.empty:
        return pd.DataFrame(columns=ROAM_COLUMNS)
    g = sub.groupby(['CdrNo','Circle'], dropna=False).agg(
        TotalCalls=('Circle','count'),
        Days=('DateStr', lambda s: s.nunique()),
        FirstLoc=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else ""),
        LastLoc=('LastCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[-1] if len(s.replace("", np.nan).dropna())>0 else ""),
        First_dt=('start_dt','min'),
        Last_dt=('start_dt','max'),
        OutCalls=('CallTypeStd', lambda s: (s=='CALL_OUT').sum()),
        InCalls=('CallTypeStd', lambda s: (s=='CALL_IN').sum()),
        OutSms=('CallTypeStd', lambda s: (s=='SMS_OUT').sum()),
        InSms=('CallTypeStd', lambda s: (s=='SMS_IN').sum()),
        TotalDuration=('Duration','sum')
    ).reset_index()
    period = g['First_dt'].dt.date.apply(fmt_date_dMonY).fillna("") + " - " + g['Last_dt'].dt.date.apply(fmt_date_dMonY).fillna("")
    df = pd.DataFrame({
        'CdrNo': g['CdrNo'],
        'Roaming': g['Circle'],
        'Period': period,
        'Total Calls': g['TotalCalls'],
        'Days': g['Days'],
        'First Location': g['FirstLoc'].apply(clean_text),
        'Last Location': g['LastLoc'].apply(clean_text),
        'Out Calls': g['OutCalls'],
        'In Calls': g['InCalls'],
        'Out Sms': g['OutSms'],
        'In Sms': g['InSms'],
        'Other Calls': 0,
        'Total Duration': g['TotalDuration']
    })
    df = df[df['Total Calls'] > 0]
    return safe_reindex_columns(df, ROAM_COLUMNS)

# 8) IMEIPeriod (unique IMEI per CdrNo)
IMEI_COLUMNS = ['CdrNo','IMEI','IMEI Manufacturer','Device Type','Period','Total Calls','Days',
 'First Location','Last Location','Out Calls','In Calls','Out Sms','In Sms','Other Calls','Total Duration']
def build_imei_period(std):
    std = std.reset_index(drop=True)
    sub = std.copy()
    sub['IMEI'] = sub['IMEI'].astype(str).str.strip()
    sub = sub[sub['IMEI'] != ""]
    if sub.empty:
        return pd.DataFrame(columns=IMEI_COLUMNS)
    g = sub.groupby(['CdrNo','IMEI'], dropna=False).agg(
        TotalCalls=('IMEI','count'),
        Days=('DateStr', lambda s: s.nunique()),
        FirstLoc=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else ""),
        LastLoc=('LastCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[-1] if len(s.replace("", np.nan).dropna())>0 else ""),
        First_dt=('start_dt','min'),
        Last_dt=('start_dt','max'),
        OutCalls=('CallTypeStd', lambda s: (s=='CALL_OUT').sum()),
        InCalls=('CallTypeStd', lambda s: (s=='CALL_IN').sum()),
        OutSms=('CallTypeStd', lambda s: (s=='SMS_OUT').sum()),
        InSms=('CallTypeStd', lambda s: (s=='SMS_IN').sum()),
        TotalDuration=('Duration','sum')
    ).reset_index()
    # Ensure uniqueness: keep highest TotalCalls then TotalDuration
    g = g.sort_values(['CdrNo','IMEI','TotalCalls','TotalDuration'], ascending=[True, True, False, False])
    g = g.drop_duplicates(['CdrNo','IMEI'], keep='first')
    period = g['First_dt'].dt.date.apply(fmt_date_dMonY).fillna("") + " - " + g['Last_dt'].dt.date.apply(fmt_date_dMonY).fillna("")
    df = pd.DataFrame({
        'CdrNo': g['CdrNo'],
        'IMEI': g['IMEI'],
        'IMEI Manufacturer': "",
        'Device Type': "",
        'Period': period,
        'Total Calls': g['TotalCalls'],
        'Days': g['Days'],
        'First Location': g['FirstLoc'].apply(clean_text),
        'Last Location': g['LastLoc'].apply(clean_text),
        'Out Calls': g['OutCalls'],
        'In Calls': g['InCalls'],
        'Out Sms': g['OutSms'],
        'In Sms': g['InSms'],
        'Other Calls': 0,
        'Total Duration': g['TotalDuration']
    })
    return safe_reindex_columns(df, IMEI_COLUMNS)

# 9) IMSIPeriod
IMSI_COLUMNS = ['CdrNo','IMSI','Period','Total Calls','Days','First Location','Last Location',
 'Out Calls','In Calls','Out Sms','In Sms','Other Calls','Total Duration']
def build_imsi_period(std):
    std = std.reset_index(drop=True)
    sub = std.copy()
    sub['IMSI'] = sub['IMSI'].astype(str).str.strip()
    sub = sub[sub['IMSI'] != ""]
    if sub.empty:
        return pd.DataFrame(columns=IMSI_COLUMNS)
    g = sub.groupby(['CdrNo','IMSI'], dropna=False).agg(
        TotalCalls=('IMSI','count'),
        Days=('DateStr', lambda s: s.nunique()),
        FirstLoc=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else ""),
        LastLoc=('LastCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[-1] if len(s.replace("", np.nan).dropna())>0 else ""),
        First_dt=('start_dt','min'),
        Last_dt=('start_dt','max'),
        OutCalls=('CallTypeStd', lambda s: (s=='CALL_OUT').sum()),
        InCalls=('CallTypeStd', lambda s: (s=='CALL_IN').sum()),
        OutSms=('CallTypeStd', lambda s: (s=='SMS_OUT').sum()),
        InSms=('CallTypeStd', lambda s: (s=='SMS_IN').sum()),
        TotalDuration=('Duration','sum')
    ).reset_index()
    period = g['First_dt'].dt.date.apply(fmt_date_dMonY).fillna("") + " - " + g['Last_dt'].dt.date.apply(fmt_date_dMonY).fillna("")
    df = pd.DataFrame({
        'CdrNo': g['CdrNo'],
        'IMSI': g['IMSI'],
        'Period': period,
        'Total Calls': g['TotalCalls'],
        'Days': g['Days'],
        'First Location': g['FirstLoc'].apply(clean_text),
        'Last Location': g['LastLoc'].apply(clean_text),
        'Out Calls': g['OutCalls'],
        'In Calls': g['InCalls'],
        'Out Sms': g['OutSms'],
        'In Sms': g['InSms'],
        'Other Calls': 0,
        'Total Duration': g['TotalDuration']
    })
    df = df[df['Total Calls'] > 0]
    return safe_reindex_columns(df, IMSI_COLUMNS)

# 10) Night_Mapping
NMAP_COLUMNS = [
 'CdrNo','B Party','Date','Time','Duration','Call Type','First Cell ID',
 'First Cell ID Address','Last Cell ID','Last Cell ID Address','IMEI',
 'IMEI Manufacturer','Device Type','IMSI','Roaming','B Party Provider',
 'Main City(First CellID)','Sub City(First CellID)','Lat-Long-Azimuth (First CellID)',
 'Crime','Circle','Operator','CallForward','LRN'
]
def build_night_mapping(std):
    std = std.reset_index(drop=True)
    sub = std[(std['IsNight'] == True)].copy()
    if sub.empty:
        return pd.DataFrame(columns=NMAP_COLUMNS)
    mapping = build_mapping(sub, include_location=False).drop(columns=['Location'], errors='ignore')
    return safe_reindex_columns(mapping, NMAP_COLUMNS)

# 11) Night_MaxStay
def build_night_maxstay(std):
    std = std.reset_index(drop=True)
    sub = std[(std['IsNight'] == True)].copy()
    if sub.empty:
        return pd.DataFrame(columns=STAY_COLUMNS)
    return build_max_stay(sub)

# 12) Day_Mapping
def build_day_mapping(std):
    std = std.reset_index(drop=True)
    sub = std[(std['IsNight'] == False)].copy()
    if sub.empty:
        return pd.DataFrame(columns=MAP_COLUMNS)
    mapping = build_mapping(sub, include_location=False)
    return safe_reindex_columns(mapping, MAP_COLUMNS)

# 13) Day_MaxStay
def build_day_maxstay(std):
    std = std.reset_index(drop=True)
    sub = std[(std['IsNight'] == False)].copy()
    if sub.empty:
        return pd.DataFrame(columns=STAY_COLUMNS)
    return build_max_stay(sub)

# 14) WorkHomeLocation
WH_COLUMNS = [
 'Location','CdrNo','Cell ID','Total Calls','Tower Address','Latitude','Longitude',
 'Azimuth','Roaming','First Call Date','First Call Time','Last Call Date','Last Call Time'
]
def build_work_home_location(std):
    std = std.reset_index(drop=True)
    rows = []
    for cdr, grp in std.groupby('CdrNo', dropna=False):
        # Home: most frequent FirstCellID during night (18–06)
        night_grp = grp[(grp['IsNight'] == True) & (grp['FirstCellID'].astype(str).str.strip() != "")]
        if not night_grp.empty:
            top = night_grp.groupby('FirstCellID', dropna=False).agg(
                TotalCalls=('FirstCellID','count'),
                First_dt=('start_dt','min'),
                Last_dt=('start_dt','max'),
                Addr=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else "")
            ).reset_index().sort_values('TotalCalls', ascending=False)
            r = top.iloc[0]
            rows.append({
                'Location': 'Home',
                'CdrNo': cdr,
                'Cell ID': r['FirstCellID'],
                'Total Calls': int(r['TotalCalls']),
                'Tower Address': clean_text(r['Addr']),
                'Latitude': "",
                'Longitude': "",
                'Azimuth': "",
                'Roaming': "",
                'First Call Date': fmt_date_dMonY(r['First_dt'].date()) if pd.notna(r['First_dt']) else "",
                'First Call Time': fmt_time_HMS(r['First_dt'].time()) if pd.notna(r['First_dt']) else "",
                'Last Call Date': fmt_date_dMonY(r['Last_dt'].date()) if pd.notna(r['Last_dt']) else "",
                'Last Call Time': fmt_time_HMS(r['Last_dt'].time()) if pd.notna(r['Last_dt']) else ""
            })
        # Work: most frequent FirstCellID during 10–18 (inclusive)
        work_grp = grp[(grp['Hour'] >= 10) & (grp['Hour'] <= 18) & (grp['FirstCellID'].astype(str).str.strip() != "")]
        if not work_grp.empty:
            top = work_grp.groupby('FirstCellID', dropna=False).agg(
                TotalCalls=('FirstCellID','count'),
                First_dt=('start_dt','min'),
                Last_dt=('start_dt','max'),
                Addr=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else "")
            ).reset_index().sort_values('TotalCalls', ascending=False)
            r = top.iloc[0]
            rows.append({
                'Location': 'Work',
                'CdrNo': cdr,
                'Cell ID': r['FirstCellID'],
                'Total Calls': int(r['TotalCalls']),
                'Tower Address': clean_text(r['Addr']),
                'Latitude': "",
                'Longitude': "",
                'Azimuth': "",
                'Roaming': "",
                'First Call Date': fmt_date_dMonY(r['First_dt'].date()) if pd.notna(r['First_dt']) else "",
                'First Call Time': fmt_time_HMS(r['First_dt'].time()) if pd.notna(r['First_dt']) else "",
                'Last Call Date': fmt_date_dMonY(r['Last_dt'].date()) if pd.notna(r['Last_dt']) else "",
                'Last Call Time': fmt_time_HMS(r['Last_dt'].time()) if pd.notna(r['Last_dt']) else ""
            })
    df = pd.DataFrame(rows, columns=WH_COLUMNS)
    return safe_reindex_columns(df, WH_COLUMNS)

# 15) HomeLocationBasedonDayFirstand
HOME2_COLUMNS = ['CdrNo','Cell ID','Total Calls','Tower Address','Latitude','Longitude','Azimuth',
 'Roaming','First Call Date','First Call Time','Last Call Date','Last Call Time']
def build_home_location_based_on_day_first(std):
    std = std.reset_index(drop=True)
    rows=[]
    for cdr, grp in std.groupby('CdrNo', dropna=False):
        night_grp = grp[(grp['IsNight'] == True) & (grp['FirstCellID'].astype(str).str.strip() != "")]
        if night_grp.empty:
            continue
        top = night_grp.groupby('FirstCellID', dropna=False).agg(
            TotalCalls=('FirstCellID','count'),
            First_dt=('start_dt','min'),
            Last_dt=('start_dt','max'),
            Addr=('FirstCellAddr', lambda s: s.replace("", np.nan).dropna().iloc[0] if len(s.replace("", np.nan).dropna())>0 else "")
        ).reset_index().sort_values('TotalCalls', ascending=False)
        r = top.iloc[0]
        rows.append({
            'CdrNo': cdr,
            'Cell ID': r['FirstCellID'],
            'Total Calls': int(r['TotalCalls']),
            'Tower Address': clean_text(r['Addr']),
            'Latitude': "",
            'Longitude': "",
            'Azimuth': "",
            'Roaming': "",
            'First Call Date': fmt_date_dMonY(r['First_dt'].date()) if pd.notna(r['First_dt']) else "",
            'First Call Time': fmt_time_HMS(r['First_dt'].time()) if pd.notna(r['First_dt']) else "",
            'Last Call Date': fmt_date_dMonY(r['Last_dt'].date()) if pd.notna(r['Last_dt']) else "",
            'Last Call Time': fmt_time_HMS(r['Last_dt'].time()) if pd.notna(r['Last_dt']) else ""
        })
    df = pd.DataFrame(rows, columns=HOME2_COLUMNS)
    return safe_reindex_columns(df, HOME2_COLUMNS)

# 16) ISDCalls (Counterparty digits length ≥ 11)
ISD_COLUMNS = [
 'CdrNo','B Party','Date','Time','Duration','Call Type','First Cell ID',
 'First Cell ID Address','Last Cell ID','Last Cell ID Address','IMEI',
 'IMEI Manufacturer','Device Type','IMSI','Roaming','B Party Country',
 'Main City(First CellID)','Sub City(First CellID)','Lat-Long-Azimuth (First CellID)',
 'Crime','Circle','Operator','LRN','CallForward'
]
def build_isd_calls(std):
    std = std.reset_index(drop=True)
    cp_digits = std['Counterparty'].astype(str).str.replace(r"\D","", regex=True)
    is_isd = cp_digits.str.len() >= 11
    sub = std[is_isd].copy()
    if sub.empty:
        return pd.DataFrame(columns=ISD_COLUMNS)
    n = len(sub)
    df = pd.DataFrame({
        'CdrNo': sub['CdrNo'],
        'B Party': sub['Counterparty'].apply(clean_text),
        'Date': sub['DateStr'],
        'Time': sub['TimeStr'],
        'Duration': sub['Duration'].fillna(0).astype('Int64'),
        'Call Type': sub['CallTypeStd'],
        'First Cell ID': sub['FirstCellID'].apply(clean_text),
        'First Cell ID Address': sub['FirstCellAddr'].apply(clean_text),
        'Last Cell ID': sub['LastCellID'].apply(clean_text),
        'Last Cell ID Address': sub['LastCellAddr'].apply(clean_text),
        'IMEI': sub['IMEI'],
        'IMEI Manufacturer': empty_series(n, ""),
        'Device Type': empty_series(n, ""),
        'IMSI': sub['IMSI'],
        'Roaming': sub['Circle'].fillna("").astype(str),
        'B Party Country': empty_series(n, ""),
        'Main City(First CellID)': sub['FirstCellCity'].apply(clean_text),
        'Sub City(First CellID)': empty_series(n, ""),
        'Lat-Long-Azimuth (First CellID)': empty_series(n, ""),
        'Crime': empty_series(n, ""),
        'Circle': sub['Circle'].fillna("").astype(str),
        'Operator': sub['operator'].fillna("").astype(str),
        'LRN': sub['LRN'].apply(clean_text),
        'CallForward': sub['CallForward'].apply(clean_text),
    })
    return safe_reindex_columns(df, ISD_COLUMNS)

# master builders
SHEET_BUILDERS = [
    ("Mapping", lambda std: build_mapping(std, include_location=False)),
    ("Summary", build_summary),
    ("MaxCalls", build_max_calls),
    ("MaxDuration", build_max_duration),
    ("MaxStay", build_max_stay),
    ("OtherStateContactSummary", build_other_state_contact_summary),
    ("RoamingPeriod", build_roaming_period),
    ("IMEIPeriod", build_imei_period),
    ("IMSIPeriod", build_imsi_period),
    ("Night_Mapping", build_night_mapping),
    ("Night_MaxStay", build_night_maxstay),
    ("Day_Mapping", build_day_mapping),
    ("Day_MaxStay", build_day_maxstay),
    ("WorkHomeLocation", build_work_home_location),
    ("HomeLocationBasedonDayFirstand", build_home_location_based_on_day_first),
    ("ISDCalls", build_isd_calls),
]

# ---------------- Excel Utilities ----------------
def trim_text_columns(df):
    """Trim spaces and collapse repeated whitespace for string-like columns."""
    out = df.copy()
    for c in out.columns:
        if out[c].dtype == "object":
            out[c] = out[c].apply(clean_text)
    return out

def autosize_sheet(ws):
    """Auto-fit column widths based on cell contents (max 60)."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            val_s = str(val)
            if len(val_s) > max_len:
                max_len = len(val_s)
        # Add small buffer; cap width
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

def postprocess_workbook(path, log_callback):
    try:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                ws.freeze_panes = "A2"
                autosize_sheet(ws)
        wb.save(path)
        log_callback("Applied AutoFilter, froze headers, and auto-sized columns on all sheets.")
    except Exception as e:
        log_callback(f"Post-formatting failed: {e}")

# ---------------- PROCESSOR ----------------
def process_files_with_selection(filepaths, output_path, selected_sheets, log_callback, progress_callback):
    all_std = []
    total = len(filepaths)
    for idx, path in enumerate(filepaths, start=1):
        log_callback(f"[{idx}/{total}] Loading: {os.path.basename(path)}")
        try:
            raw = load_csv_like_raw(path)
        except Exception as e:
            log_callback(f"Failed to load {path}: {e}")
            continue
        log_callback(f"   → Raw rows: {len(raw)} | cols: {len(raw.columns)}")
        try:
            std = standardize_rows(raw)
        except Exception as e:
            log_callback(f"   → Standardize failed: {e}")
            continue
        log_callback(f"   → Standardized rows: {len(std)}")
        all_std.append(std)
        progress_callback(idx/total)
    if not all_std:
        log_callback("No valid CSVs loaded. Aborting.")
        return
    merged = pd.concat(all_std, ignore_index=True)
    merged = merged.sort_values("start_dt").reset_index(drop=True)
    log_callback(f"Merged total rows: {len(merged)}")

    # Build only selected sheets
    log_callback("Building selected sheets...")
    sheets = {}
    selected_builders = [(name, func) for name, func in SHEET_BUILDERS if name in selected_sheets]
    
    for name, func in selected_builders:
        try:
            df = func(merged)
            if isinstance(df, pd.Series):
                df = df.to_frame().T
            # Trim spaces in text columns to avoid visual gaps, keep numeric ints
            df = trim_text_columns(df)
            sheets[name] = df
            log_callback(f"  ✓ {name}: rows={len(df)} cols={len(df.columns)}")
        except Exception as e:
            log_callback(f"  ✗ {name} build failed: {e}")
            sheets[name] = pd.DataFrame()  # empty fallback

    # Write to Excel
    log_callback(f"Writing Excel to {output_path} ...")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            # safe sheet name length
            sheet_name = name[:31]
            try:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            except Exception as e:
                log_callback(f"Write failed for {name}: {e}")

    # Post-format: AutoFilter, Freeze, Auto-size
    postprocess_workbook(output_path, log_callback)
    log_callback("All done.")
    progress_callback(1.0)

def process_files(filepaths, output_path, log_callback, progress_callback):
    all_std = []
    total = len(filepaths)
    for idx, path in enumerate(filepaths, start=1):
        log_callback(f"[{idx}/{total}] Loading: {os.path.basename(path)}")
        try:
            raw = load_csv_like_raw(path)
        except Exception as e:
            log_callback(f"Failed to load {path}: {e}")
            continue
        log_callback(f"   → Raw rows: {len(raw)} | cols: {len(raw.columns)}")
        try:
            std = standardize_rows(raw)
        except Exception as e:
            log_callback(f"   → Standardize failed: {e}")
            continue
        log_callback(f"   → Standardized rows: {len(std)}")
        all_std.append(std)
        progress_callback(idx/total)
    if not all_std:
        log_callback("No valid CSVs loaded. Aborting.")
        return
    merged = pd.concat(all_std, ignore_index=True)
    merged = merged.sort_values("start_dt").reset_index(drop=True)
    log_callback(f"Merged total rows: {len(merged)}")

    # Build sheets
    log_callback("Building sheets...")
    sheets = {}
    for name, func in SHEET_BUILDERS:
        try:
            df = func(merged)
            if isinstance(df, pd.Series):
                df = df.to_frame().T
            # Trim spaces in text columns to avoid visual gaps, keep numeric ints
            df = trim_text_columns(df)
            sheets[name] = df
            log_callback(f"  ✓ {name}: rows={len(df)} cols={len(df.columns)}")
        except Exception as e:
            log_callback(f"  ✗ {name} build failed: {e}")
            sheets[name] = pd.DataFrame()  # empty fallback

    # Write to Excel
    log_callback(f"Writing Excel to {output_path} ...")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            # safe sheet name length
            sheet_name = name[:31]
            try:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            except Exception as e:
                log_callback(f"Write failed for {name}: {e}")

    # Post-format: AutoFilter, Freeze, Auto-size
    postprocess_workbook(output_path, log_callback)
    log_callback("All done.")
    progress_callback(1.0)

# ---------------- GUI ----------------
class CDRApp:
    def __init__(self, root):
        self.root = root
        root.title("CDR Analyzer — Enhanced")
        root.geometry("1200x800")
        self.filepaths = []
        self.output_path = ""
        
        # Sheet selection variables
        self.sheet_vars = {}
        sheet_names = [name for name, _ in SHEET_BUILDERS]
        for sheet_name in sheet_names:
            self.sheet_vars[sheet_name] = tk.BooleanVar(value=True)  # All selected by default
        
        self._setup_style()
        self._build_ui()
        self.log("Ready. Select files and Save As before Analyze.")

    def _setup_style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure(".", font=("Segoe UI", 10))
        style.configure("Card.TFrame", background="#F3F4F6")
        style.configure("LabelMuted.TLabel", background="#F3F4F6", foreground="#6B7280", font=("Segoe UI", 9, "bold"))
        style.configure("Blue.TButton", background="#3B82F6", foreground="#ffffff", padding=8)
        style.configure("Green.TButton", background="#10B981", foreground="#ffffff", padding=8)
        style.configure("Red.TButton", background="#EF4444", foreground="#ffffff", padding=8)
        style.configure("Orange.TButton", background="#F59E0B", foreground="#ffffff", padding=8)
        style.configure("Purple.TButton", background="#8B5CF6", foreground="#ffffff", padding=8)
        style.configure("Thin.Horizontal.TProgressbar", troughcolor="#E5E7EB", background="#3B82F6", thickness=12)
        style.configure("Section.TLabel", background="#F3F4F6", foreground="#374151", font=("Segoe UI", 10, "bold"))
        style.configure("Body.TLabel", background="#F3F4F6", foreground="#4B5563", font=("Segoe UI", 9))

    def _build_ui(self):
        # Main container with two columns
        main_container = ttk.Frame(self.root)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Left column for controls and sheet selection
        left_frame = ttk.Frame(main_container, style="Card.TFrame", padding=10)
        left_frame.pack(side="left", fill="y", padx=(0,5))
        
        # Right column for log and progress
        right_frame = ttk.Frame(main_container)
        right_frame.pack(side="right", fill="both", expand=True, padx=(5,0))
        
        # === LEFT COLUMN ===
        # Control buttons
        ttk.Label(left_frame, text="📋 CDR Analysis Controls", style="Section.TLabel").pack(anchor="w", pady=(0,10))
        
        self.btn_browse = ttk.Button(left_frame, text="📂 Browse CSV Files", style="Blue.TButton", command=self.browse_files)
        self.btn_browse.pack(fill="x", pady=2)
        
        self.btn_save = ttk.Button(left_frame, text="💾 Save Location", style="Blue.TButton", command=self.save_as)
        self.btn_save.pack(fill="x", pady=2)
        
        # File selection status
        self.sel_label = ttk.Label(left_frame, text="No files selected.", style="Body.TLabel")
        self.sel_label.pack(anchor="w", pady=(10,5))
        
        # Sheet selection section
        ttk.Label(left_frame, text="📊 Select Sheets to Generate", style="Section.TLabel").pack(anchor="w", pady=(20,10))
        
        # Sheet selection controls
        select_frame = ttk.Frame(left_frame)
        select_frame.pack(fill="x", pady=(0,10))
        
        self.btn_select_all = ttk.Button(select_frame, text="Select All", style="Purple.TButton", command=self.select_all_sheets)
        self.btn_select_all.pack(side="left", padx=(0,5))
        
        self.btn_select_none = ttk.Button(select_frame, text="Select None", style="Orange.TButton", command=self.select_no_sheets)
        self.btn_select_none.pack(side="left")
        
        # Scrollable frame for checkboxes
        canvas = tk.Canvas(left_frame, height=300, bg="#F3F4F6", highlightthickness=0)
        scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, style="Card.TFrame")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Create checkboxes for each sheet
        for i, (sheet_name, _) in enumerate(SHEET_BUILDERS):
            cb = ttk.Checkbutton(scrollable_frame, text=sheet_name, variable=self.sheet_vars[sheet_name])
            cb.pack(anchor="w", pady=1, padx=5)
        
        # Action buttons
        ttk.Label(left_frame, text="⚡ Actions", style="Section.TLabel").pack(anchor="w", pady=(20,10))
        
        self.btn_analyze = ttk.Button(left_frame, text="▶ Start Analysis", style="Green.TButton", command=self.analyze)
        self.btn_analyze.pack(fill="x", pady=2)
        
        self.btn_clear = ttk.Button(left_frame, text="🧹 Clear Log", style="Orange.TButton", command=self.clear_log)
        self.btn_clear.pack(fill="x", pady=2)
        
        self.btn_exit = ttk.Button(left_frame, text="⏻ Exit", style="Red.TButton", command=self.root.quit)
        self.btn_exit.pack(fill="x", pady=2)
        
        # === RIGHT COLUMN ===
        # Progress section
        prog_wrap = ttk.Frame(right_frame, style="Card.TFrame", padding=10)
        prog_wrap.pack(fill="x", pady=(0,10))
        ttk.Label(prog_wrap, text="📈 Progress", style="Section.TLabel").pack(anchor="w", pady=(0,5))
        self.progress = ttk.Progressbar(prog_wrap, orient="horizontal", mode="determinate", style="Thin.Horizontal.TProgressbar")
        self.progress.pack(fill="x")
        
        # Log section
        log_card = ttk.Frame(right_frame, style="Card.TFrame", padding=10)
        log_card.pack(fill="both", expand=True)
        ttk.Label(log_card, text="📝 Analysis Log", style="Section.TLabel").pack(anchor="w", pady=(0,5))
        self.logbox = scrolledtext.ScrolledText(log_card, wrap="word", font=("Consolas",9), bg="#111827", fg="#e5e7eb", insertbackground="#e5e7eb")
        self.logbox.pack(fill="both", expand=True)

    def log(self, msg):
        self.logbox.insert("end", msg+"\n")
        self.logbox.see("end")
        self.root.update_idletasks()

    def clear_log(self):
        self.logbox.delete("1.0","end")
    
    def select_all_sheets(self):
        for var in self.sheet_vars.values():
            var.set(True)
        self.log("Selected all sheets for generation.")
    
    def select_no_sheets(self):
        for var in self.sheet_vars.values():
            var.set(False)
        self.log("Deselected all sheets.")

    def browse_files(self):
        files = filedialog.askopenfilenames(title="Select CDR CSV files", filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if not files:
            return
        self.filepaths = list(files)
        self.sel_label.config(text=f"{len(self.filepaths)} file(s) selected.")
        self.log("Selected files:")
        for f in self.filepaths:
            self.log("  - " + f)

    def save_as(self):
        path = filedialog.asksaveasfilename(title="Save output Excel", defaultextension=".xlsx", filetypes=[("Excel workbook","*.xlsx")])
        if not path:
            return
        self.output_path = path
        self.log(f"Output will be saved to: {path}")

    def analyze(self):
        if not self.filepaths:
            messagebox.showerror("Missing input", "Please select at least one CSV file.")
            return
        if not self.output_path:
            messagebox.showerror("Missing output", "Please choose a save location for the Excel.")
            return
        
        # Get selected sheets
        selected_sheets = [name for name, var in self.sheet_vars.items() if var.get()]
        if not selected_sheets:
            messagebox.showerror("No sheets selected", "Please select at least one sheet to generate.")
            return
        
        try:
            self.progress["value"] = 0
            self.log("Starting analysis...")
            self.log(f"Selected sheets: {', '.join(selected_sheets)}")
            process_files_with_selection(self.filepaths, self.output_path, selected_sheets, self.log, self._progress_cb)
            self.log("Analysis finished.")
            messagebox.showinfo("Done", f"Analysis complete. Saved: {self.output_path}")
        except Exception as e:
            self.log(f"ERROR: {e}")
            messagebox.showerror("Analysis error", str(e))

    def _progress_cb(self, val):
        try:
            self.progress["value"] = val * 100.0
            self.root.update_idletasks()
        except Exception:
            pass

# ---------------- MAIN ----------------
if __name__ == "__main__":
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    app = CDRApp(root)
    root.mainloop()