 #!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Generator v2 (final patched, fixed indentation)
- Creates exactly 9 sheets (_01 … _09) as requested
- Highlights important headers
- Fix: safe handling of NaT/NaN dates in Relationship and Mobile SwitchOFF sheets
- Backward compatibility: keep generate_excel_file()
"""

import re
import logging
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    def __init__(self, progress_callback=None):
        self.progress_callback = progress_callback
        self.cancel_flag = False

    def set_cancel_flag(self):
        self.cancel_flag = True

    def update_progress(self, percent, message=""):
        if self.progress_callback:
            self.progress_callback(percent, message)

    def clean_text(self, s):
        if s is None or (isinstance(s, float) and np.isnan(s)):
            return ""
        return str(s).strip()

    # -------------------------
    # Autofit and Styling
    # -------------------------
    def autofit_and_style(self, workbook, sheet_name, important_headers, sheet_index=0):
        try:
            ws = workbook[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column

            # --- Tab colors palette ---
            tab_colors = [
                "92D050", "4472C4", "ED7D31", "7030A0", "C00000",
                "00B0F0", "FFC000", "548235", "2E75B6"
            ]
            ws.sheet_properties.tabColor = tab_colors[sheet_index % len(tab_colors)]

            # Freeze top row and first column
            ws.freeze_panes = "B2"

            # --- Styles ---
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            imp_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")   # dark blue
            normal_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")  # lighter blue
            alt_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
            highlight_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")  # yellow
            thin = Side(border_style="thin", color="FF999999")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # --- Headers ---
            important_indexes = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                header_value = str(cell.value) if cell.value else ""

                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

                if header_value in important_headers:
                    cell.fill = imp_fill
                    important_indexes.append(col_idx)
                else:
                    cell.fill = normal_fill

            # --- Highlight important columns ---
            for r in range(2, max_row + 1):
                for col_idx in important_indexes:
                    ws.cell(row=r, column=col_idx).fill = highlight_fill

            # --- Alternating row fill for non-important cols ---
            for r in range(2, max_row + 1):
                if r % 2 == 0:
                    for c in range(1, max_col + 1):
                        if c not in important_indexes:
                            ws.cell(row=r, column=c).fill = alt_fill

            # --- Autofit columns ---
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    v = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(v))
                ws.column_dimensions[col_letter].width = min(50, max(10, max_length + 3))

        except Exception as e:
            logging.warning(f"Styling error on sheet {sheet_name}: {e}")

    # -------------------------
    # Utility
    # -------------------------
    def drop_empty_rows(self, df):
        if df is None or df.empty:
            return df
        return df.replace(r"^\s*$", np.nan, regex=True).dropna(how="all")

    # -------------------------
    # Sheet creators (all kept as-is)
    # -------------------------
    def create__01_CDR_Format(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "CDR Party No", "Opposite Party No", "Opp Party-Name", "Opp Party-Full Address", "Opp Party-SP State",
                "CALL_DATE", "CALL_TIME", "Call_Type_Std", "CALL_DURATION", "FIRST_CELL_ID_A", "First_Cell_Site_Address",
                "First_Cell_Site_Name-City", "First_Lat_Long", "LAST_CELL_ID_A", "Last_Cell_Site_Address",
                "Last_Cell_Site_Name-City", "Last_Lat_Long", "ESN_IMEI_A", "IMSI_A", "CUST_TYPE", "SMSC_CENTER",
                "Home Circle", "ROAM_CIRCLE", "Opp Party-Activation Date", "Opp Party-Service Provider", "ID"
            ])
        out = df[[
            "CDR Party No", "Opposite Party No", "Opp Party-Name", "Opp Party-Full Address", "Opp Party-SP State",
            "CALL_DATE", "CALL_TIME", "CallTypeStd", "CALL_DURATION", "FIRST_CELL_ID_A", "First_Cell_Site_Address",
            "First_Cell_Site_Name-City", "First_Lat_Long", "LAST_CELL_ID_A", "Last_Cell_Site_Address",
            "Last_Cell_Site_Name-City", "Last_Lat_Long", "ESN_IMEI_A", "IMSI_A", "CUST_TYPE", "SMSC_CENTER",
            "Home Circle", "ROAM_CIRCLE", "Opp Party-Activation Date", "Opp Party-Service Provider", "ID"
        ]].copy()
        return out.rename(columns={"CallTypeStd": "Call_Type_Std"})

    def create__02_Relationship_Call_Frequ(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "ID", "CDR Party No", "Opposite Party No", "Opp Party-SP State",
                "Opp Party-Name", "Opp Party-Full Address", "Start_Date", "End_Date",
                "Date_Diff", "Total Event", "Call In", "Call Out", "SMS In", "SMS Out",
                "Call In_Duration", "Call Out_Duration", "Total_Duration"
            ])

        df["date_only"] = pd.to_datetime(df["CALL_DATE"], errors="coerce").dt.date
        grp = df.groupby(["CDR Party No", "Opposite Party No"], dropna=False)

        rows = []
        idx = 1

        for (cdr, opp), g in grp:
            if str(opp).strip() == "":
                continue

            # Only use valid dates
            valid_dates = [d for d in g["date_only"] if isinstance(d, date) and pd.notna(d)]
            if valid_dates:
                start = min(valid_dates)
                end = max(valid_dates)
                date_diff = (end - start).days
            else:
                start = None
                end = None
                date_diff = 0

            rows.append({
                "ID": idx,
                "CDR Party No": cdr,
                "Opposite Party No": opp,
                "Opp Party-SP State": g["ROAM_CIRCLE"].mode().iat[0]
                    if not g["ROAM_CIRCLE"].replace("", np.nan).dropna().empty else "",
                "Opp Party-Name": g["Opp Party-Name"].iloc[0]
                    if "Opp Party-Name" in g.columns else str(opp),
                "Opp Party-Full Address": "",
                "Start_Date": start.strftime("%Y-%m-%d") if (start and pd.notna(start) and isinstance(start, date)) else "",
                "End_Date": end.strftime("%Y-%m-%d") if (end and pd.notna(end) and isinstance(end, date)) else "",
                "Date_Diff": date_diff,
                "Total Event": len(g),
                "Call In": int((g["CallTypeStd"] == "CALL_IN").sum()),
                "Call Out": int((g["CallTypeStd"] == "CALL_OUT").sum()),
                "SMS In": int((g["CallTypeStd"] == "SMS_IN").sum()),
                "SMS Out": int((g["CallTypeStd"] == "SMS_OUT").sum()),
                "Call In_Duration": int(g.loc[g["CallTypeStd"] == "CALL_IN", "DurationSeconds"].sum()),
                "Call Out_Duration": int(g.loc[g["CallTypeStd"] == "CALL_OUT", "DurationSeconds"].sum()),
                "Total_Duration": int(g["DurationSeconds"].sum())
            })
            idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Total Event", "Total_Duration"], ascending=[False, False]
        ).reset_index(drop=True)


    def create__03_Cell_ID_Frequency(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "Id", "CDR Party No", "FIRST_CELL_ID_A", "First_Cell_Site_Address",
                "First_Lat_Long", "Total Event", "Call In", "Call Out", "SMS In", "SMS Out",
                "Call In_Duration", "Call Out_Duration", "Total_Duration", "ROAM_CIRCLE",
                "First_Cell_Site_Name-City"
            ])

        grp = df.groupby(["CDR Party No", "FIRST_CELL_ID_A"], dropna=False)
        rows = []
        idx = 1

        for (cdr, cell), g in grp:
            if str(cell).strip() == "":
                continue

            rows.append({
                "Id": idx,
                "CDR Party No": cdr,
                "FIRST_CELL_ID_A": cell,
                "First_Cell_Site_Address": g["First_Cell_Site_Address"].mode().iat[0]
                    if not g["First_Cell_Site_Address"].replace("", np.nan).dropna().empty else "",
                "First_Lat_Long": g["First_Lat_Long"].mode().iat[0]
                    if not g["First_Lat_Long"].replace("", np.nan).dropna().empty else "",
                "Total Event": len(g),
                "Call In": int((g["CallTypeStd"] == "CALL_IN").sum()),
                "Call Out": int((g["CallTypeStd"] == "CALL_OUT").sum()),
                "SMS In": int((g["CallTypeStd"] == "SMS_IN").sum()),
                "SMS Out": int((g["CallTypeStd"] == "SMS_OUT").sum()),
                "Call In_Duration": int(g.loc[g["CallTypeStd"] == "CALL_IN", "DurationSeconds"].sum()),
                "Call Out_Duration": int(g.loc[g["CallTypeStd"] == "CALL_OUT", "DurationSeconds"].sum()),
                "Total_Duration": int(g["DurationSeconds"].sum()),
                "ROAM_CIRCLE": g["ROAM_CIRCLE"].mode().iat[0]
                    if not g["ROAM_CIRCLE"].replace("", np.nan).dropna().empty else "",
                "First_Cell_Site_Name-City": g["First_Cell_Site_Name-City"].mode().iat[0]
                    if not g["First_Cell_Site_Name-City"].replace("", np.nan).dropna().empty else ""
            })
            idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Total Event"], ascending=False
        ).reset_index(drop=True)


    def create__04_Movement_Analysis(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "ID", "CDR Party No", "Opposite Party No", "CALL_DATE", "CALL_TIME",
                "FIRST_CELL_ID_A", "First_Cell_Site_Name-City",
                "First_Cell_Site_Address", "First_Lat_Long"
            ])

        out = df[[
            "CDR Party No", "Opposite Party No", "CALL_DATE", "CALL_TIME",
            "FIRST_CELL_ID_A", "First_Cell_Site_Name-City",
            "First_Cell_Site_Address", "First_Lat_Long"
        ]].copy()

        out.insert(0, "ID", range(1, len(out) + 1))
        
        out["CALL_DATETIME"] = pd.to_datetime(
            out["CALL_DATE"].astype(str) + " " + out["CALL_TIME"].astype(str),
            errors="coerce"
        )
        return out.sort_values(by=["CALL_DATETIME"]).drop(columns=["CALL_DATETIME"]).reset_index(drop=True)

    def create__05_Imei_Used(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "ID", "CDR Party No", "CDR Party-Name", "CDR Party-Full Address",
                "CDR Party-Service Provider", "IMEI", "First_Call", "Last_call",
                "Total Event", "Call In", "Call Out", "SMS In", "SMS Out",
                "Call In_Duration", "Call Out_Duration", "Total_Duration"
            ])

        grp = df[df["ESN_IMEI_A"].astype(str).str.strip() != ""].groupby(
            ["CDR Party No", "ESN_IMEI_A"], dropna=False
        )

        rows = []
        idx = 1

        for (cdr, imei), g in grp:
            first_dt = g["start_dt"].min()
            last_dt = g["start_dt"].max()

            rows.append({
                "ID": idx,
                "CDR Party No": cdr,
                "CDR Party-Name": "",
                "CDR Party-Full Address": "",
                "CDR Party-Service Provider": g["Opp Party-Service Provider"].mode().iat[0]
                    if "Opp Party-Service Provider" in g.columns and not g["Opp Party-Service Provider"].replace("", np.nan).dropna().empty else "",
                "IMEI": imei,
                "First_Call": first_dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(first_dt) else "",
                "Last_call": last_dt.strftime("%Y-%m-%d %H:%M:%S") if pd.notna(last_dt) else "",
                "Total Event": len(g),
                "Call In": int((g["CallTypeStd"] == "CALL_IN").sum()),
                "Call Out": int((g["CallTypeStd"] == "CALL_OUT").sum()),
                "SMS In": int((g["CallTypeStd"] == "SMS_IN").sum()),
                "SMS Out": int((g["CallTypeStd"] == "SMS_OUT").sum()),
                "Call In_Duration": int(g.loc[g["CallTypeStd"] == "CALL_IN", "DurationSeconds"].sum()),
                "Call Out_Duration": int(g.loc[g["CallTypeStd"] == "CALL_OUT", "DurationSeconds"].sum()),
                "Total_Duration": int(g["DurationSeconds"].sum())
            })
            idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Total Event"], ascending=False
        ).reset_index(drop=True)

    def create__06_State_Connection(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "Id", "CDR Party No", "Connection of State", "Total Event",
                "Call In", "Call Out", "SMS In", "SMS Out",
                "Call In_Duration", "Call Out_Duration", "Total_Duration"
            ])

        df["ConnectionState"] = (
            df["Home Circle"].fillna("")
            .replace("", np.nan)
            .fillna(df["ROAM_CIRCLE"].fillna(""))
        )
        grp = df.groupby(["CDR Party No", "ConnectionState"], dropna=False)

        rows = []
        idx = 1

        for (cdr, state), g in grp:
            if str(state).strip() == "":
                continue

            rows.append({
                "Id": idx,
                "CDR Party No": cdr,
                "Connection of State": state,
                "Total Event": len(g),
                "Call In": int((g["CallTypeStd"] == "CALL_IN").sum()),
                "Call Out": int((g["CallTypeStd"] == "CALL_OUT").sum()),
                "SMS In": int((g["CallTypeStd"] == "SMS_IN").sum()),
                "SMS Out": int((g["CallTypeStd"] == "SMS_OUT").sum()),
                "Call In_Duration": int(g.loc[g["CallTypeStd"] == "CALL_IN", "DurationSeconds"].sum()),
                "Call Out_Duration": int(g.loc[g["CallTypeStd"] == "CALL_OUT", "DurationSeconds"].sum()),
                "Total_Duration": int(g["DurationSeconds"].sum())
            })
            idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Total Event"], ascending=False
        ).reset_index(drop=True)
        

    def create__07_ISD_Call(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "CdrNo", "B Party", "Date", "Time", "Duration", "Call Type",
                "First Cell ID", "First Cell ID Address", "Last Cell ID",
                "Last Cell ID Address", "IMEI", "IMSI", "Roaming", "Operator"
            ])

        def is_international(number):
            if pd.isna(number) or str(number).strip() == "":
                return False
            s = str(number).strip()
            num_clean = re.sub(r"\D", "", s)
            return s.startswith("+") or s.startswith("00") or (len(num_clean) > 12)

        calls = df[df["CallTypeStd"].str.startswith("CALL")].copy()
        calls["IsISD"] = calls["Opposite Party No"].apply(is_international)
        isd = calls[calls["IsISD"]]

        if isd.empty:
            return pd.DataFrame(columns=[
                "CdrNo", "B Party", "Date", "Time", "Duration", "Call Type",
                "First Cell ID", "First Cell ID Address", "Last Cell ID",
                "Last Cell ID Address", "IMEI", "IMSI", "Roaming", "Operator"
            ])

        out = isd[[
            "CDR Party No", "Opposite Party No", "CALL_DATE", "CALL_TIME",
            "CALL_DURATION", "FIRST_CELL_ID_A", "First_Cell_Site_Address",
            "LAST_CELL_ID_A", "Last_Cell_Site_Address", "ESN_IMEI_A", "IMSI_A",
            "ROAM_CIRCLE", "Opp Party-Service Provider"
        ]].copy()

        out = out.rename(columns={
            "Opposite Party No": "B Party",
            "CALL_DATE": "Date",
            "CALL_TIME": "Time",
            "CALL_DURATION": "Duration",
            "FIRST_CELL_ID_A": "First Cell ID",
            "First_Cell_Site_Address": "First Cell ID Address",
            "LAST_CELL_ID_A": "Last Cell ID",
            "Last_Cell_Site_Address": "Last Cell ID Address",
            "ESN_IMEI_A": "IMEI",
            "IMSI_A": "IMSI",
            "ROAM_CIRCLE": "Roaming",
            "Opp Party-Service Provider": "Operator"
        })

        out.insert(0, "CdrNo", out.pop("CDR Party No"))
        out["Call Type"] = isd["CallTypeStd"].values

        out["DATETIME"] = pd.to_datetime(
            out["Date"].astype(str) + " " + out["Time"].astype(str),
            errors="coerce"
        )
        return out.sort_values(by=["DATETIME"]).drop(columns=["DATETIME"]).reset_index(drop=True)


    def create__08_Night_Call(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "Id", "CDR Party No", "Opposite Party No", "Opp Party-Name",
                "Opp Party-Full Address", "Opp Party-SP State", "Total Event",
                "Call In", "Call Out", "SMS In", "SMS Out",
                "Call In_Duration", "Call Out_Duration", "Total_Duration"
            ])

        night = df[df["IsNight"] == True].copy()
        grp = night.groupby(["CDR Party No", "Opposite Party No"], dropna=False)

        rows = []
        idx = 1

        for (cdr, opp), g in grp:
            if str(opp).strip() == "":
                continue

            rows.append({
                "Id": idx,
                "CDR Party No": cdr,
                "Opposite Party No": opp,
                "Opp Party-Name": g["Opp Party-Name"].iloc[0]
                    if "Opp Party-Name" in g.columns else str(opp),
                "Opp Party-Full Address": "",
                "Opp Party-SP State": g["ROAM_CIRCLE"].mode().iat[0]
                    if not g["ROAM_CIRCLE"].replace("", np.nan).dropna().empty else "",
                "Total Event": len(g),
                "Call In": int((g["CallTypeStd"] == "CALL_IN").sum()),
                "Call Out": int((g["CallTypeStd"] == "CALL_OUT").sum()),
                "SMS In": int((g["CallTypeStd"] == "SMS_IN").sum()),
                "SMS Out": int((g["CallTypeStd"] == "SMS_OUT").sum()),
                "Call In_Duration": int(g.loc[g["CallTypeStd"] == "CALL_IN", "DurationSeconds"].sum()),
                "Call Out_Duration": int(g.loc[g["CallTypeStd"] == "CALL_OUT", "DurationSeconds"].sum()),
                "Total_Duration": int(g["DurationSeconds"].sum())
            })
            idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Total Event"], ascending=False
        ).reset_index(drop=True)


    def create__09_Mobile_SwitchOFF(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=["ID", "Start_Date", "End_Date", "Total_Day"])

        df["date_only"] = pd.to_datetime(df["CALL_DATE"], errors="coerce").dt.date
        grp = df.groupby(["CDR Party No"], dropna=False)
        rows = []
        idx = 1

        for cdr, g in grp:
            # Only use valid dates
            dates = [d for d in g["date_only"] if isinstance(d, date) and pd.notna(d)]
            if not dates:
                continue

            dates = sorted(set(dates))
            for i in range(len(dates) - 1):
                gap = (dates[i + 1] - dates[i]).days
                if gap > 1:
                    rows.append({
                        "ID": idx,
                        "Start_Date": dates[i].strftime("%Y-%m-%d")
                            if (dates[i] and pd.notna(dates[i]) and isinstance(dates[i], date)) else "",
                        "End_Date": dates[i + 1].strftime("%Y-%m-%d")
                            if (dates[i + 1] and pd.notna(dates[i + 1]) and isinstance(dates[i + 1], date)) else "",
                        "Total_Day": gap
                    })
                    idx += 1

        return pd.DataFrame(rows).sort_values(
            by=["Start_Date"], ascending=True
        ).reset_index(drop=True)


    # -------------------------
    # Main generate function
    # -------------------------
    def generate_excel(self, df, output_path):
        try:
            self.update_progress(40, "Generating Excel...")

            sheet_defs = [
                ("_01_CDR_Format", self.create__01_CDR_Format,
                 ["CDR Party No", "Opposite Party No", "CALL_DATE", "CALL_TIME", "CALL_DURATION"]),
                ("_02_Relationship_Call_Frequ", self.create__02_Relationship_Call_Frequ,
                 ["CDR Party No", "Opposite Party No", "Start_Date", "End_Date", "Total Event"]),
                ("_03_Cell_ID_Frequency", self.create__03_Cell_ID_Frequency,
                 ["CDR Party No", "FIRST_CELL_ID_A", "Total Event"]),
                ("_04_Movement_Analysis", self.create__04_Movement_Analysis,
                 ["CDR Party No", "Opposite Party No", "CALL_DATE", "FIRST_CELL_ID_A"]),
                ("_05_Imei_Used", self.create__05_Imei_Used,
                 ["CDR Party No", "IMEI", "First_Call", "Last_call", "Total Event"]),
                ("_06_State_Connection", self.create__06_State_Connection,
                 ["CDR Party No", "Connection of State", "Total Event"]),
                ("_07_ISD_Call", self.create__07_ISD_Call,
                 ["CdrNo", "B Party", "Date", "Duration"]),
                ("_08_Night_Call", self.create__08_Night_Call,
                 ["CDR Party No", "Opposite Party No", "Total Event"]),
                ("_09_Mobile_SwitchOFF", self.create__09_Mobile_SwitchOFF,
                 ["Start_Date", "End_Date", "Total_Day"])
            ]

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, creator, imp_cols in sheet_defs:
                    if self.cancel_flag:
                        raise Exception("Cancelled")
                    self.update_progress(50, f"Generating {sheet_name}")
                    sheet_df = creator(df)
                    sheet_df = self.drop_empty_rows(sheet_df)
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # styling pass
            wb = load_workbook(output_path)
            for idx, (sheet_name, _, imp_cols) in enumerate(sheet_defs):
                self.autofit_and_style(wb, sheet_name, imp_cols, sheet_index=idx)
            wb.save(output_path)

            self.update_progress(100, f"Excel generated: {output_path}")
            return output_path

        except Exception as e:
            logging.error(f"Error generating Excel: {e}")
            raise

    # -------------------------
    # Backward compatibility
    # -------------------------
    def generate_excel_file(self, df, output_path):
        """Legacy wrapper for compatibility with old modules."""
        return self.generate_excel(df, output_path)
